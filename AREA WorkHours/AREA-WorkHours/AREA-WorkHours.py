import csv
import datetime as dt
import os
import re
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

DEPT_ORDER_KEYWORDS = [
    "七工",
    "台中工地", "台中設計", "台中",
    "嘉義", "屏東", "高雄", "內湖", "萬大", "大潭",
]

OUTPUT_HEADERS = [
    "日期", "姓名", "員工編號", "部門", "班別",
    "上班", "刷卡", "下班", "刷卡", "備註",
    "遲到/P", "事假/H", "病假/H", "請假時間", "出差",
    "加班", "前2/H", "後6/H", "加班區間", "超/H", "合計", "加倍/H", "週日改獎勵金", "出工津貼/次",
]

PRESERVE_COLS = [
    "備註", "遲到/P", "事假/H", "病假/H", "請假時間", "出差",
    "加班", "前2/H", "後6/H", "加班區間", "超/H", "合計", "加倍/H", "週日改獎勵金", "出工津貼/次",
]

HOURS_COLS = {
    "前2/H", "後6/H", "超/H", "合計", "加倍/H",
    "事假/H", "病假/H",
}
INT_COLS = {
    "遲到/P", "週日改獎勵金", "出工津貼/次",
}

GREEN_HEADER_NAMES = {
    "事假/H", "病假/H", "請假時間", "出差",
    "加班", "前2/H", "後6/H", "加班區間", "超/H", "合計", "加倍/H", "週日改獎勵金", "出工津貼/次",
}

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_ROW_FILL = PatternFill("solid", fgColor="FFC7CE")

HEADER_FONT = Font(bold=True)
GREEN_ROW_FONT = Font(bold=True)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

THIN = Side(style="thin", color="D9D9D9")
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FMT = "dd/mm/yyyy"
TIME_FMT = "hh:mm"
HOURS_FMT = "0.0"
INT_FMT = "0"

RED_RGB = "FFFF0000"
MAX_AUTOFIT_WIDTH = 160

YMD_RE = re.compile(r"^\s*(\d{4})[/-](\d{1,2})[/-](\d{1,2})\s*$")
DMY_RE = re.compile(r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{4})\s*$")
HM_RE = re.compile(r"^\s*(\d{1,2}):(\d{2})\s*$")

DT1_RE = re.compile(r"(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2})(?::\d{2})?")
DT2_RE = re.compile(r"(\d{4}/\d{2}/\d{2})\s+(\d{2}:\d{2})(?::\d{2})?")

OT_APPROVED_VALUES = {"已加班", "簽核中"}
OT_BLOCK_MINUTES = 30

DT_ANY_RE = re.compile(r"(\d{4}[-/]\d{2}[-/]\d{2})\s+(\d{2}:\d{2})(?::\d{2})?")


def cell_str(v) -> str:
    return "" if v is None else str(v)


def pick_file():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    return filedialog.askopenfilename(
        title="Select the raw CSV export",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    ) or None


def clean_text(s) -> str:
    if s is None:
        return ""
    s = str(s).replace("\r", " ").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def dept_rank(dept_text: str) -> int:
    dept_text = dept_text or ""
    for i, kw in enumerate(DEPT_ORDER_KEYWORDS):
        if kw and kw in dept_text:
            return i
    return len(DEPT_ORDER_KEYWORDS)


def uniquify_headers(headers):
    counts = defaultdict(int)
    out = []
    for h in headers:
        h = (h or "").strip()
        counts[h] += 1
        out.append(h if counts[h] == 1 else f"{h}__{counts[h]}")
    return out


def infer_dayfirst(date_strings):
    for s in date_strings:
        m = DMY_RE.match(s or "")
        if not m:
            continue
        a = int(m.group(1))
        b = int(m.group(2))
        if a > 12 and b <= 12:
            return True
        if b > 12 and a <= 12:
            return False
    return True


def parse_date(value, *, dayfirst=True):
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        s = value.strip()
        m = YMD_RE.match(s)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return dt.date(y, mo, d)
            except Exception:
                return None
        m = DMY_RE.match(s)
        if m:
            a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return dt.date(y, b, a) if dayfirst else dt.date(y, a, b)
            except Exception:
                return None
    return None


def parse_time(value):
    if value in (None, ""):
        return None
    if isinstance(value, dt.time):
        return value
    if isinstance(value, dt.datetime):
        return value.time()
    if isinstance(value, str):
        m = HM_RE.match(value.strip())
        if m:
            hh, mm = int(m.group(1)), int(m.group(2))
            if 0 <= hh <= 23 and 0 <= mm <= 59:
                return dt.time(hh, mm)
    return None


def infer_schedule_from_shift(shift_text: str):
    if not shift_text:
        return None, None
    s = str(shift_text)
    if "不用打卡" in s:
        return None, None
    if "8點半" in s or "8:30" in s or "08:30" in s:
        return dt.time(8, 30), dt.time(17, 30)
    if "9點半" in s or "9:30" in s or "09:30" in s:
        return dt.time(9, 30), dt.time(18, 30)
    if "8點" in s:
        return dt.time(8, 0), dt.time(17, 0)
    if "9點" in s:
        return dt.time(9, 0), dt.time(18, 0)
    return None, None


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def to_float(v):
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("，", ",")
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        if "," in s and "." not in s:
            if re.match(r"^\d+,\d+$", s):
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def floor_to_half_hour(hours: float) -> float:
    return (int(hours * 2)) / 2.0


def fmt_hhmm(t: dt.time) -> str:
    return f"{t.hour:02d}:{t.minute:02d}"


def make_font_red(cell):
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic, underline=f.underline,
        color=Color(rgb=RED_RGB),
    )


def make_substring_red(cell, keyword: str):
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
    except Exception:
        make_font_red(cell)
        return

    text = "" if cell.value is None else str(cell.value)
    if not text or keyword not in text:
        return

    parts = text.split(keyword)
    rt = CellRichText()
    normal = InlineFont()
    red = InlineFont(color=Color(rgb=RED_RGB))

    for i, p in enumerate(parts):
        if p:
            rt.append(TextBlock(normal, p))
        if i < len(parts) - 1:
            rt.append(TextBlock(red, keyword))

    cell.value = rt


def parse_dt_token(date_s: str, time_s: str):
    y, m, d = map(int, date_s.replace("/", "-").split("-"))
    hh, mm = map(int, time_s.split(":"))
    return dt.datetime(y, m, d, hh, mm)


def parse_dt_range(text: str):
    if not text:
        return None, None
    s = str(text).strip()

    m = DT1_RE.findall(s)
    if not m:
        m = DT2_RE.findall(s)
    if len(m) >= 2:
        start = parse_dt_token(m[0][0], m[0][1])
        end = parse_dt_token(m[1][0], m[1][1])
        return start, end
    return None, None


def normalize_leave_type(s: str) -> str:
    s = clean_text(s)
    if "病假" in s:
        return "病假"
    if "事假" in s:
        return "事假"
    if "特休" in s:
        return "特休"
    return s


def looks_like_leave_type(s: str) -> bool:
    s = clean_text(s)
    if not s:
        return False
    return ("假" in s) or (s in {"特休", "公假", "喪假", "婚假", "產假", "陪產", "補休"})


def ceil_to_block(t: dt.datetime, minutes=OT_BLOCK_MINUTES) -> dt.datetime:
    m = t.minute
    add = (minutes - (m % minutes)) % minutes
    if add == 0 and t.second == 0 and t.microsecond == 0:
        return t
    return (t.replace(second=0, microsecond=0) + dt.timedelta(minutes=add))


def floor_to_block(t: dt.datetime, minutes=OT_BLOCK_MINUTES) -> dt.datetime:
    m = t.minute - (t.minute % minutes)
    return t.replace(minute=m, second=0, microsecond=0)


def diff_minutes(a: dt.datetime, b: dt.datetime) -> int:
    return int((b - a).total_seconds() // 60)


def approved_ot_value(v: str) -> str:
    v = clean_text(v)
    if v in OT_APPROVED_VALUES:
        return v
    return ""


def rec_get_first(rec: dict, base_key: str) -> str:
    if base_key in rec and clean_text(rec.get(base_key)) != "":
        return clean_text(rec.get(base_key))
    for k, v in rec.items():
        if k == base_key or k.startswith(base_key + "__"):
            vv = clean_text(v)
            if vv != "":
                return vv
    return ""


def parse_raw_ot_pairs(raw_text: str):
    s = clean_text(raw_text)
    if not s:
        return []

    tokens = []
    for m in DT_ANY_RE.finditer(s):
        date_s = m.group(1).replace("/", "-")
        time_s = m.group(2)
        y, mo, d = map(int, date_s.split("-"))
        hh, mm = map(int, time_s.split(":"))
        tokens.append(dt.datetime(y, mo, d, hh, mm))

    pairs = []
    for i in range(0, len(tokens) - 1, 2):
        a, b = tokens[i], tokens[i + 1]
        if b > a:
            pairs.append((a, b))
    return pairs


def normalize_ot_segments(pairs):
    segs = []
    for s, e in pairs:
        s2 = ceil_to_block(s)
        e2 = floor_to_block(e)
        if e2 > s2:
            segs.append((s2, e2))
    return segs


def split_front2_after6(total_hours: float):
    f2 = min(total_hours, 2.0)
    a6 = max(total_hours - 2.0, 0.0)
    a6 = min(a6, 6.0)
    over = max(total_hours - (2.0 + 6.0), 0.0)
    return f2, a6, over


def format_ot_range_text_from_segments(segs):
    if not segs:
        return "", 0.0, 0.0, 0.0, 0.0

    total_minutes = sum(diff_minutes(s, e) for s, e in segs)
    total_hours = floor_to_half_hour(total_minutes / 60.0)
    if total_hours <= 0:
        return "", 0.0, 0.0, 0.0, 0.0

    f2, a6, over = split_front2_after6(total_hours)

    rem_f2 = int(round(f2 * 60))
    rem_a6 = int(round(a6 * 60))
    rem_over = int(round(over * 60))

    parts = []
    for s, e in segs:
        cur = s
        while cur < e:
            nxt = min(cur + dt.timedelta(minutes=OT_BLOCK_MINUTES), e)
            block_min = diff_minutes(cur, nxt)

            if rem_f2 > 0:
                label = "前2"
                take = min(block_min, rem_f2)
                rem_f2 -= take
            elif rem_a6 > 0:
                label = "後6"
                take = min(block_min, rem_a6)
                rem_a6 -= take
            elif rem_over > 0:
                label = "超"
                take = min(block_min, rem_over)
                rem_over -= take
            else:
                label = "未計"

            if parts and parts[-1]["label"] == label and parts[-1]["end"] == cur:
                parts[-1]["end"] = nxt
            else:
                parts.append({"label": label, "start": cur, "end": nxt})

            cur = nxt

    text_parts = [
        f"{fmt_hhmm(p['start'].time())}-{fmt_hhmm(p['end'].time())}({p['label']})"
        for p in parts
        if p["end"] > p["start"]
    ]

    return " + ".join(text_parts), f2, a6, over, (f2 + a6 + over)


def style_header_row(ws):
    for col_idx, name in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = name
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = GRID_BORDER
        if name in GREEN_HEADER_NAMES:
            cell.fill = GREEN_FILL
    ws.freeze_panes = "A2"


def style_cell(cell, *, align="center", is_date=False, is_time=False, is_hours=False, is_int=False):
    cell.border = GRID_BORDER
    cell.alignment = CENTER if align == "center" else LEFT
    if is_date and cell.value:
        cell.number_format = DATE_FMT
    if is_time and cell.value:
        cell.number_format = TIME_FMT
    if is_hours and cell.value not in (None, ""):
        cell.number_format = HOURS_FMT
    if is_int and cell.value not in (None, ""):
        cell.number_format = INT_FMT


def apply_green_row_style(ws, row_idx, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row_idx, c)
        cell.fill = GREEN_FILL
        cell.font = GREEN_ROW_FONT
        cell.alignment = CENTER
        cell.border = GRID_BORDER


def apply_red_row_style(ws, row_idx, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row_idx, c)
        cell.fill = RED_ROW_FILL
        cell.border = GRID_BORDER


def autofit_columns(ws, col_letters):
    for col in col_letters:
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws[f"{col}{r}"].value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        width = min(MAX_AUTOFIT_WIDTH, max(10, max_len + 2))
        ws.column_dimensions[col].width = width


def sniff_delimiter(sample_text: str) -> str:
    for d in [",", ";", "\t"]:
        if sample_text.count(d) >= 5:
            return d
    try:
        return csv.Sniffer().sniff(sample_text, delimiters=[",", ";", "\t"]).delimiter
    except Exception:
        return ","


def read_csv_rows(csv_path: str):
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        raw = f.read()

    delim = sniff_delimiter(raw[:5000])
    rows = list(csv.reader(raw.splitlines(), delimiter=delim, quotechar='"'))

    repaired = []
    for row in rows:
        if len(row) == 1 and isinstance(row[0], str) and ("," in row[0] or ";" in row[0] or "\t" in row[0]):
            d2 = sniff_delimiter(row[0])
            repaired.append(next(csv.reader([row[0]], delimiter=d2, quotechar='"')))
        else:
            repaired.append(row)

    repaired = [r for r in repaired if any((c or "").strip() for c in r)]
    return repaired


def build_rows_from_records(source_headers, records):
    has_swipe1 = "刷卡" in source_headers
    has_swipe2 = "刷卡__2" in source_headers

    date_strs = []
    for rec in records:
        s = clean_text(rec.get("日期") or "")
        if DMY_RE.match(s):
            date_strs.append(s)
    dayfirst = infer_dayfirst(date_strs)

    rows = []
    for rec in records:
        r = {h: "" for h in OUTPUT_HEADERS}

        r["日期"] = clean_text(rec.get("日期") or "")
        r["姓名"] = clean_text(rec.get("姓名") or "")
        r["員工編號"] = clean_text(rec.get("員工編號") or "")
        r["部門"] = clean_text(rec.get("部門") or "")
        r["班別"] = clean_text(rec.get("班別") or "")

        r["上班"] = clean_text(rec.get("上班") or "")
        r["下班"] = clean_text(rec.get("下班") or "")

        r["刷卡#1"] = clean_text(rec.get("刷卡") or "") if has_swipe1 else ""
        r["刷卡#2"] = clean_text(rec.get("刷卡__2") or "") if has_swipe2 else ""

        leave_time = clean_text(rec.get("請假時間") or "")
        leave_type = clean_text(rec.get("請假假別") or "")

        trip_flag = clean_text(rec.get("出差") or "")
        trip_time = clean_text(rec.get("出差時間") or "")

        if leave_time:
            r["請假時間"] = leave_time

        if trip_flag and trip_time:
            r["出差"] = f"{trip_flag} {trip_time}"
        elif trip_time:
            r["出差"] = trip_time
        elif trip_flag:
            r["出差"] = trip_flag

        ot_status = approved_ot_value(rec_get_first(rec, "加班"))
        if ot_status:
            r["加班"] = ot_status

        raw_ot_time = rec_get_first(rec, "加班時間")
        if ot_status and raw_ot_time:
            pairs = parse_raw_ot_pairs(raw_ot_time)
            segs = normalize_ot_segments(pairs)
            ot_text, f2, a6, over, total = format_ot_range_text_from_segments(segs)

            if ot_text:
                r["加班區間"] = ot_text
            if f2 > 0:
                r["前2/H"] = float(f2)
            if a6 > 0:
                r["後6/H"] = float(a6)
            if over > 0:
                r["超/H"] = float(over)
            if total > 0:
                r["合計"] = float(total)

        note1 = clean_text(rec.get("備註") or "")
        note2 = clean_text(rec.get("備註__2") or "")
        desc = clean_text(rec.get("說明") or "")
        combined_note = " ".join([x for x in [note1, note2, desc] if x]).strip()

        if leave_type and looks_like_leave_type(leave_type):
            combined_note = (combined_note + " " + leave_type).strip() if combined_note else leave_type

        if combined_note:
            r["備註"] = combined_note

        for col in PRESERVE_COLS:
            if col in ("備註", "請假時間", "出差", "加班區間", "加班"):
                continue
            if col not in rec:
                continue

            raw = clean_text(rec.get(col) or "")
            if raw == "":
                continue

            if col in HOURS_COLS:
                r[col] = to_float(raw)
            elif col in INT_COLS:
                r[col] = int(round(to_float(raw)))
            else:
                r[col] = raw

        rows.append(r)

    return rows, dayfirst


def sort_and_group(rows, *, dayfirst):
    def emp_key(r):
        return (r.get("員工編號") or "").strip() or (r.get("姓名") or "").strip()

    def date_key(r):
        d = parse_date(r.get("日期"), dayfirst=dayfirst)
        return d or dt.date.max

    def dept_key(r):
        dept = (r.get("部門") or "").strip()
        return (dept_rank(dept), dept)

    rows_sorted = sorted(rows, key=lambda r: (dept_key(r), emp_key(r), date_key(r)))

    final = []
    prev_emp = None
    prev_dept = None

    for r in rows_sorted:
        cur_emp = emp_key(r)
        cur_dept = (r.get("部門") or "").strip()

        if prev_emp is None:
            prev_emp, prev_dept = cur_emp, cur_dept

        if cur_emp != prev_emp or cur_dept != prev_dept:
            final.append(None)
            prev_emp, prev_dept = cur_emp, cur_dept

        final.append(r)

    if final:
        final.append(None)

    return final


def write_formatted_sheet(ws, final_rows, *, dayfirst):
    ws.delete_rows(1, ws.max_row)
    style_header_row(ws)

    def set_cell_value(cell, v):
        if v is None:
            cell.value = None
        elif isinstance(v, str) and v.strip() == "":
            cell.value = None
        else:
            cell.value = v

    max_col = len(OUTPUT_HEADERS)
    out_r = 2
    sep_rows = []

    for item in final_rows:
        if item is None:
            for c in range(1, max_col + 1):
                ws.cell(out_r, c).value = None
            apply_green_row_style(ws, out_r, max_col)
            sep_rows.append(out_r)
            out_r += 1
            continue

        r = item
        for col_idx, h in enumerate(OUTPUT_HEADERS, start=1):
            cell = ws.cell(out_r, col_idx)

            if h == "刷卡":
                if col_idx == 7:
                    val = r.get("刷卡#1", "")
                elif col_idx == 9:
                    val = r.get("刷卡#2", "")
                else:
                    val = ""
            else:
                val = r.get(h, "")

            if h == "日期":
                d = parse_date(val, dayfirst=dayfirst)
                cell.value = d if d else (val if (isinstance(val, str) and val.strip()) else None)
                style_cell(cell, align="center", is_date=bool(d))

            elif col_idx in (6, 7, 8, 9):
                t = parse_time(val)
                cell.value = t if t else (val if (isinstance(val, str) and val.strip()) else None)
                style_cell(cell, align="center", is_time=bool(t))

            elif h in ("姓名", "部門", "班別", "備註", "請假時間", "出差", "加班區間", "加班"):
                set_cell_value(cell, val)
                style_cell(cell, align="left")

            else:
                set_cell_value(cell, val)
                if h in HOURS_COLS:
                    style_cell(cell, align="center", is_hours=True)
                elif h in INT_COLS:
                    style_cell(cell, align="center", is_int=True)
                else:
                    style_cell(cell, align="center")

        out_r += 1

    ws.column_dimensions["A"].width = 14
    for col in ("F", "G", "H", "I"):
        ws.column_dimensions[col].width = 8

    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    cols = []
    for name in ("備註", "請假時間", "出差", "加班區間"):
        c = header_to_col.get(name)
        if c:
            cols.append(get_column_letter(c))
    autofit_columns(ws, cols)

    return sep_rows


def compute_and_fill(ws, sep_rows):
    header_map = {}
    swipe_cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v == "刷卡":
            swipe_cols.append(c)
        if isinstance(v, str) and v.strip():
            header_map[v.strip()] = c
    if len(swipe_cols) < 2:
        return

    col_date = header_map["日期"]
    col_shift = header_map["班別"]
    col_sched_start = header_map["上班"]
    col_sched_end = header_map["下班"]

    col_swipe_a = swipe_cols[0]
    col_swipe_b = swipe_cols[1]

    col_note = header_map["備註"]
    col_late = header_map["遲到/P"]
    col_pers = header_map["事假/H"]
    col_sick = header_map["病假/H"]
    col_leave_time = header_map["請假時間"]
    col_trip = header_map["出差"]

    col_ot_status = header_map["加班"]

    max_col = len(OUTPUT_HEADERS)
    sep_set = set(sep_rows)

    samples = []
    for r in range(2, min(ws.max_row, 200) + 1):
        v = ws.cell(r, col_date).value
        if isinstance(v, str) and (YMD_RE.match(v.strip()) or DMY_RE.match(v.strip())):
            samples.append(v.strip())
    dayfirst = infer_dayfirst(samples)

    invalid_rows = set()

    def cell_blank(rr, cc):
        return is_blank(ws.cell(rr, cc).value)

    def append_note(rr, text, make_red=False):
        note_cell = ws.cell(rr, col_note)
        current = cell_str(note_cell.value).strip()
        if text in current:
            return
        note_cell.value = (current + " " + text).strip() if current else text
        if make_red:
            make_font_red(note_cell)

    for r in range(2, ws.max_row + 1):
        if r in sep_set:
            continue

        raw_date = ws.cell(r, col_date).value
        d = raw_date if isinstance(raw_date, dt.date) else parse_date(raw_date, dayfirst=dayfirst)
        if not d:
            if any(not is_blank(ws.cell(r, c).value) for c in range(1, max_col + 1)):
                apply_red_row_style(ws, r, max_col)
                invalid_rows.add(r)
            continue

        sched_start = parse_time(ws.cell(r, col_sched_start).value)
        sched_end = parse_time(ws.cell(r, col_sched_end).value)
        if sched_start is None or sched_end is None:
            s_start, s_end = infer_schedule_from_shift(ws.cell(r, col_shift).value)
            sched_start = sched_start or s_start
            sched_end = sched_end or s_end

        t1 = parse_time(ws.cell(r, col_swipe_a).value)
        t2 = parse_time(ws.cell(r, col_swipe_b).value)
        times = [t for t in (t1, t2) if t is not None]
        swipe_in = min(times) if times else None
        swipe_out = max(times) if times else None

        swipe_in_cell = None
        swipe_out_cell = None
        if swipe_in is not None:
            swipe_in_cell = ws.cell(r, col_swipe_a) if t1 == swipe_in else ws.cell(r, col_swipe_b)
        if swipe_out is not None:
            swipe_out_cell = ws.cell(r, col_swipe_b) if t2 == swipe_out else ws.cell(r, col_swipe_a)

        pers_h_existing = to_float(ws.cell(r, col_pers).value)
        sick_h_existing = to_float(ws.cell(r, col_sick).value)
        has_leave = (not is_blank(ws.cell(r, col_leave_time).value)) or (pers_h_existing > 0) or (sick_h_existing > 0)
        has_trip = not is_blank(ws.cell(r, col_trip).value)

        shift_text = clean_text(ws.cell(r, col_shift).value or "")
        no_clock = "不用打卡" in shift_text

        note_text = clean_text(cell_str(ws.cell(r, col_note).value))
        missing_swipe_flag = ("未刷卡" in note_text) or ("忘刷" in note_text) or ("下班未刷卡" in note_text)

        for kw in ("未刷卡", "忘刷", "下班未刷卡"):
            if kw in note_text:
                make_substring_red(ws.cell(r, col_note), kw)

        ot_status = approved_ot_value(ws.cell(r, col_ot_status).value)
        ot_allowed = ot_status in OT_APPROVED_VALUES

        if (not has_leave) and (not has_trip) and (not no_clock):
            if swipe_in and swipe_out:
                if swipe_in == swipe_out:
                    if not missing_swipe_flag:
                        apply_red_row_style(ws, r, max_col)
                        invalid_rows.add(r)
                        continue
                dt_in = dt.datetime.combine(d, swipe_in)
                dt_out = dt.datetime.combine(d, swipe_out)
                if dt_out <= dt_in:
                    apply_red_row_style(ws, r, max_col)
                    invalid_rows.add(r)
                    continue

            if (swipe_in is None) != (swipe_out is None):
                if not missing_swipe_flag:
                    apply_red_row_style(ws, r, max_col)
                    invalid_rows.add(r)
                    continue

            if (swipe_in or swipe_out) and (sched_start is None or sched_end is None) and (not ot_allowed):
                apply_red_row_style(ws, r, max_col)
                invalid_rows.add(r)
                continue

        if r in invalid_rows:
            continue

        leave_text = ws.cell(r, col_leave_time).value
        if (not is_blank(leave_text)) and (not has_trip):
            start_dt, end_dt = parse_dt_range(str(leave_text))
            if start_dt and end_dt and end_dt > start_dt and sched_start and sched_end:
                work_start = dt.datetime.combine(d, sched_start)
                work_end = dt.datetime.combine(d, sched_end)
                overlap_start = max(start_dt, work_start)
                overlap_end = min(end_dt, work_end)
                leave_minutes = int((overlap_end - overlap_start).total_seconds() // 60) if overlap_end > overlap_start else 0
                leave_hours = floor_to_half_hour(leave_minutes / 60.0)

                if leave_hours > 0:
                    leave_type = normalize_leave_type(note_text)
                    if "病假" in leave_type and cell_blank(r, col_sick):
                        ws.cell(r, col_sick).value = float(leave_hours)
                        ws.cell(r, col_sick).number_format = HOURS_FMT
                    elif "事假" in leave_type and cell_blank(r, col_pers):
                        ws.cell(r, col_pers).value = float(leave_hours)
                        ws.cell(r, col_pers).number_format = HOURS_FMT
                    else:
                        if looks_like_leave_type(leave_type):
                            tag = f"{leave_type} {leave_hours:.1f}H"
                            if tag and tag not in note_text:
                                ws.cell(r, col_note).value = (note_text + " " + tag).strip() if note_text else tag

        schedule_missing = (sched_start is None) or (sched_end is None) or (
            is_blank(ws.cell(r, col_sched_start).value) and is_blank(ws.cell(r, col_sched_end).value)
        )
        skip_late_early = has_leave or has_trip or (ot_allowed and schedule_missing)

        if (not skip_late_early) and sched_start and swipe_in:
            t_sched = dt.datetime.combine(d, sched_start)
            t_in = dt.datetime.combine(d, swipe_in)
            diff = int((t_in - t_sched).total_seconds() // 60)
            if diff > 0 and cell_blank(r, col_late):
                ws.cell(r, col_late).value = int(diff)
                ws.cell(r, col_late).number_format = INT_FMT
                if swipe_in_cell:
                    make_font_red(swipe_in_cell)
                make_font_red(ws.cell(r, col_late))
                append_note(r, "遲到", make_red=True)

        if (not skip_late_early) and sched_end and swipe_out:
            t_sched_end = dt.datetime.combine(d, sched_end)
            t_out = dt.datetime.combine(d, swipe_out)
            early = int((t_sched_end - t_out).total_seconds() // 60)
            if early > 0:
                if swipe_out_cell:
                    make_font_red(swipe_out_cell)
                append_note(r, "早退", make_red=True)

                cur = int(round(to_float(ws.cell(r, col_late).value))) if not cell_blank(r, col_late) else 0
                ws.cell(r, col_late).value = int(cur + early)
                ws.cell(r, col_late).number_format = INT_FMT
                make_font_red(ws.cell(r, col_late))

    def set_sum_formula(row_idx: int, col_idx: int, start_row: int, end_row: int, decimals: int):
        cell = ws.cell(row_idx, col_idx)
        col_letter = get_column_letter(col_idx)
        if end_row < start_row:
            cell.value = 0
        else:
            rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
            cell.value = f"=ROUND(SUM({rng}),{decimals})"

    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    numeric_cols = []
    for name in (list(HOURS_COLS) + list(INT_COLS)):
        if name in header_to_col:
            numeric_cols.append((name, header_to_col[name]))

    start = 2
    for sep_r in sorted(sep_rows):
        end = sep_r - 1

        for name, col_idx in numeric_cols:
            if name in INT_COLS:
                set_sum_formula(sep_r, col_idx, start, end, decimals=0)
                ws.cell(sep_r, col_idx).number_format = INT_FMT
            else:
                set_sum_formula(sep_r, col_idx, start, end, decimals=1)
                ws.cell(sep_r, col_idx).number_format = HOURS_FMT

        apply_green_row_style(ws, sep_r, max_col)
        start = sep_r + 1


def process_csv_file(csv_path: str) -> str:
    rows = read_csv_rows(csv_path)
    if not rows:
        raise ValueError("CSV is empty.")

    source_headers = uniquify_headers(rows[0])
    records = []
    for row in rows[1:]:
        if len(row) < len(source_headers):
            row += [""] * (len(source_headers) - len(row))
        elif len(row) > len(source_headers):
            row = row[:len(source_headers)]
        records.append({source_headers[i]: row[i] for i in range(len(source_headers))})

    out_rows, dayfirst = build_rows_from_records(source_headers, records)
    final = sort_and_group(out_rows, dayfirst=dayfirst)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EIP刷卡紀錄"

    sep_rows = write_formatted_sheet(ws, final, dayfirst=dayfirst)
    compute_and_fill(ws, sep_rows)

    out_dir = os.path.dirname(csv_path)
    out_path = os.path.join(out_dir, "115_月考勤計薪_請假加班紀錄.xlsx")
    wb.save(out_path)
    return out_path


def main():
    import sys
    path = sys.argv[1] if len(sys.argv) >= 2 else pick_file()
    if not path:
        print("No file selected.")
        return

    if os.path.splitext(path)[1].lower() != ".csv":
        print("Please select the raw .csv export file.")
        return

    out = process_csv_file(path)
    print("Saved:", out)


if __name__ == "__main__":
    main()
