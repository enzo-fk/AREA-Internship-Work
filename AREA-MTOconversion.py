# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
import math
import csv
from dataclasses import dataclass
from collections import defaultdict
from typing import Dict, List, Tuple, Optional

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def norm_text(x: object) -> str:
    if x is None:
        return ""
    return re.sub(r"\s+", " ", str(x)).strip()

def norm_text_lc(x: object) -> str:
    return norm_text(x).lower()

def parse_number(x: object) -> Optional[float]:
    """Parse numbers that may use comma decimals (e.g., 0,5840)."""
    if x is None:
        return None
    s = str(x).strip()
    if s == "" or s == "**":
        return None
    s = s.replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(".") > s.rfind(","):
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
    else:
        if "," in s and "." not in s:
            s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

def parse_yes(v: object) -> bool:
    s = norm_text_lc(v)
    return s in {"yes", "y", "true", "1", "是", "有", "v"}

def round_up_100(n: int) -> int:
    if n <= 0:
        return 0
    return int(math.ceil(n / 100.0) * 100)


def f_total_weight(row_idx: int) -> str:
    return f'=IF(OR(NOT(ISNUMBER(H{row_idx})),NOT(ISNUMBER(G{row_idx}))),"**",H{row_idx}*G{row_idx})'

def f_total_surface(row_idx: int) -> str:
    return f'=IF(OR(NOT(ISNUMBER(J{row_idx})),NOT(ISNUMBER(G{row_idx}))),"**",J{row_idx}*G{row_idx})'


_INCH_TOKEN_RE = re.compile(r"-\s*([^-\(\)]+?)\s*B\b", re.IGNORECASE)

def inch_token_to_float(token: str) -> float:
    token = norm_text(token)
    if token == "":
        return float("nan")
    parts = token.split()
    total = 0.0
    for part in parts:
        if "/" in part:
            try:
                a, b = part.split("/", 1)
                total += float(a) / float(b)
            except Exception:
                pass
        else:
            try:
                total += float(part)
            except Exception:
                pass
    return total

def parse_inch_token(type_str: str) -> Tuple[str, float]:
    s = norm_text(type_str)
    m = _INCH_TOKEN_RE.search(s)
    if not m:
        return ("", float("nan"))
    token = re.sub(r"\s+", " ", norm_text(m.group(1)))
    return token, inch_token_to_float(token)

def inch_band(inch_value: float) -> str:
    if math.isnan(inch_value):
        return "UNKNOWN"
    if inch_value < 10:
        return "LT10"
    if 10 <= inch_value <= 14:
        return "10_14"
    return "GE16"

def contains_inch_text(text_lc: str, inch_token_text: str) -> bool:
    if not text_lc or text_lc == "**":
        return False
    tok = norm_text_lc(inch_token_text)
    if tok == "":
        return False
    tok_re = re.escape(tok).replace(r"\ ", r"\s+")
    pattern = re.compile(rf"(?<!\d){tok_re}\s*[\"″”]", re.IGNORECASE)
    return bool(pattern.search(text_lc))

def parse_inch_range_from_text(text_lc: str) -> Optional[Tuple[float, float]]:
    if not text_lc:
        return None
    m = re.search(r"<\s*([0-9]+(?:\s+[0-9]+/[0-9]+|/[0-9]+)?)\s*[\"″”]", text_lc)
    if m:
        hi = inch_token_to_float(m.group(1))
        return (-1e9, hi)
    tokens = re.findall(r"([0-9]+(?:\s+[0-9]+/[0-9]+|/[0-9]+)?)\s*[\"″”]", text_lc)
    if len(tokens) >= 2:
        lo = inch_token_to_float(tokens[0])
        hi = inch_token_to_float(tokens[1])
        if not math.isnan(lo) and not math.isnan(hi):
            return (min(lo, hi), max(lo, hi))
    return None


@dataclass(frozen=True)
class MasterRec:
    item_no: str
    material: str
    name: str
    size: str
    treatment: str
    unit: str
    unit_weight: Optional[float]
    unit_surface: Optional[float]
    remark: str
    add_notes: str
    add_notes_lc: str

def canon_header(h: object) -> str:
    s = norm_text(h).lower()
    s = s.replace("\n", " ")
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

def map_master_columns(headers: List[object]) -> Dict[str, int]:
    """
    Returns mapping for keys:
      itemno, material, name, size, treatment, unit, qty, unitweight, unitsurface, remark, addnotes
    """
    idx = {canon_header(h): i+1 for i, h in enumerate(headers) if norm_text(h) != ""}
    def find_col(*tokens: str) -> Optional[int]:
        for k, c in idx.items():
            if all(t in k for t in tokens):
                return c
        return None

    out = {}
    out["itemno"] = find_col("item", "no") or find_col("itemno")
    out["material"] = find_col("material")
    out["name"] = find_col("name")
    out["size"] = find_col("size")
    out["treatment"] = find_col("treatment")
    out["unit"] = find_col("unit")
    out["qty"] = find_col("qty") or find_col("qtyty") or find_col("qty")
    out["unitweight"] = find_col("unit", "weight")
    out["unitsurface"] = find_col("unit", "surface")
    out["remark"] = find_col("remark")
    out["addnotes"] = find_col("add", "notes") or find_col("addnotes")

    essential = ["material", "name", "size"]
    for e in essential:
        if not out.get(e):
            raise ValueError(f"Master.xlsx missing required column for {e}. Headers found: {headers}")

    return out

def load_master(master_path: str) -> Tuple[Dict[str, MasterRec], List[MasterRec]]:
    wb = openpyxl.load_workbook(master_path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col = map_master_columns(headers)

    def get(r: int, key: str) -> object:
        c = col.get(key)
        if not c:
            return None
        return ws.cell(row=r, column=c).value

    recs: List[MasterRec] = []
    by_size: Dict[str, MasterRec] = {}

    for r in range(2, ws.max_row + 1):
        size = norm_text(get(r, "size"))
        name = norm_text(get(r, "name"))
        material = norm_text(get(r, "material"))
        if size == "" and name == "" and material == "":
            continue

        rec = MasterRec(
            item_no=norm_text(get(r, "itemno")) or "**",
            material=material or "**",
            name=name or "**",
            size=size or "**",
            treatment=norm_text(get(r, "treatment")) or "**",
            unit=norm_text(get(r, "unit")) or "**",
            unit_weight=parse_number(get(r, "unitweight")),
            unit_surface=parse_number(get(r, "unitsurface")),
            remark=norm_text(get(r, "remark")) or "**",
            add_notes=norm_text(get(r, "addnotes")) or "**",
            add_notes_lc=norm_text_lc(get(r, "addnotes")),
        )
        recs.append(rec)
        k = norm_text_lc(rec.size)
        if k and k not in by_size:
            by_size[k] = rec

    return by_size, recs

def find_first(recs: List[MasterRec], pred) -> Optional[MasterRec]:
    for r in recs:
        if pred(r):
            return r
    return None

def find_all(recs: List[MasterRec], pred) -> List[MasterRec]:
    out = []
    for r in recs:
        if pred(r):
            out.append(r)
    return out


def pick_padding_plate(master_recs: List[MasterRec], inch_token_text: str) -> Optional[MasterRec]:
    inch_lc = norm_text_lc(inch_token_text)
    return find_first(
        master_recs,
        lambda r: r.name.lower() == "plate"
        and "type 52&66" in r.add_notes_lc
        and "padding" in r.add_notes_lc
        and contains_inch_text(r.add_notes_lc, inch_lc)
    )

def pick_small_reinforcement_plate(master_recs: List[MasterRec]) -> Optional[MasterRec]:
    return find_first(master_recs, lambda r: r.name.lower() == "plate" and "small reinforcement plate" in r.add_notes_lc)

def pick_reinforcement_plate(master_recs: List[MasterRec], inch_token_text: str) -> Optional[MasterRec]:
    inch_lc = norm_text_lc(inch_token_text)
    return find_first(
        master_recs,
        lambda r: r.name.lower() == "plate"
        and "reinforcement plate" in r.add_notes_lc
        and contains_inch_text(r.add_notes_lc, inch_lc)
    )

def pick_pipe_shoe_plates(master_recs: List[MasterRec], inch_value: float) -> List[MasterRec]:
    out = []
    for r in master_recs:
        if r.name.lower() != "plate":
            continue
        if "pipe shoe material" not in r.add_notes_lc:
            continue
        if "type 52&66" not in r.add_notes_lc:
            continue
        rng = parse_inch_range_from_text(r.add_notes_lc)
        if rng and rng[0] <= inch_value <= rng[1]:
            out.append(r)
    return out

def pick_h_channel(master_recs: List[MasterRec], inch_value: float) -> Optional[MasterRec]:
    best = None
    best_span = None
    for r in master_recs:
        if r.name.lower() != "h channel":
            continue
        rng = parse_inch_range_from_text(r.add_notes_lc)
        if not rng:
            continue
        if rng[0] <= inch_value <= rng[1]:
            span = rng[1] - rng[0]
            if best is None or span < (best_span or 1e18):
                best = r
                best_span = span
    return best

def pick_forming_angle(master_by_size: Dict[str, MasterRec], length_mm: int) -> Optional[MasterRec]:
    size = f"L 40x40x5x{length_mm}"
    return master_by_size.get(norm_text_lc(size))

def pick_pipe_clamp(master_recs: List[MasterRec], inch_token_text: str) -> Optional[MasterRec]:
    inch_lc = norm_text_lc(inch_token_text)
    return find_first(
        master_recs,
        lambda r: r.name.lower() == "pipe clamp"
        and "type 54a,54b" in r.add_notes_lc
        and contains_inch_text(r.add_notes_lc, inch_lc)
    )

def pick_gasket(master_recs: List[MasterRec], inch_token_text: str) -> Optional[MasterRec]:
    inch_lc = norm_text_lc(inch_token_text)
    return find_first(
        master_recs,
        lambda r: r.name.lower() == "non-asbestos compressed gasket"
        and "type 54a,54b" in r.add_notes_lc
        and contains_inch_text(r.add_notes_lc, inch_lc)
    )

def pick_hex_bolt_set(master_recs: List[MasterRec], inch_value: float) -> Optional[MasterRec]:
    candidates = []
    for r in master_recs:
        if r.name.lower() != "hex  bolt set":
            continue
        rng = parse_inch_range_from_text(r.add_notes_lc)
        if rng:
            if rng[0] <= inch_value <= rng[1]:
                candidates.append((rng[1] - rng[0], r))
            continue
        toks = re.findall(r"([0-9]+(?:\s+[0-9]+/[0-9]+|/[0-9]+)?)\s*[\"″”]", r.add_notes_lc)
        if toks:
            exact = inch_token_to_float(toks[0])
            if not math.isnan(exact) and abs(exact - inch_value) < 1e-9:
                candidates.append((0.0, r))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]


def sniff_sep(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        return dialect.delimiter
    except Exception:
        return "\t" if sample.count("\t") >= sample.count(",") else ","

def read_delimited_flexible(path: str, sep: str) -> pd.DataFrame:
    rows: List[List[str]] = []
    with open(path, "r", encoding="utf-8-sig", errors="ignore", newline="") as f:
        reader = csv.reader(f, delimiter=sep)
        for row in reader:
            rows.append([cell for cell in row])

    if not rows:
        return pd.DataFrame()

    max_len = max(len(r) for r in rows)
    for r in rows:
        if len(r) < max_len:
            r.extend([""] * (max_len - len(r)))

    return pd.DataFrame(rows)

def read_input_table(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()

    if ext in [".xlsx", ".xlsm", ".xls"]:
        raw = pd.read_excel(path, header=None, dtype=str)
        raw = raw.fillna("")
    else:
        with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
            sample = f.read(4096)

        seps = []
        seps.append(sniff_sep(sample))
        for s in ["\t", ",", ";", "|"]:
            if s not in seps:
                seps.append(s)

        best_raw = None
        best_score = -1

        def score_raw(raw_df: pd.DataFrame) -> int:
            df0 = raw_df.fillna("")
            df0 = df0.loc[:, (df0.astype(str).apply(lambda col: (col.str.strip() != "").any())).values]
            if df0.shape[1] == 0 or df0.shape[0] == 0:
                return -1

            header_row = None
            for i in range(min(len(df0), 40)):
                cells = [norm_text_lc(x) for x in df0.iloc[i].tolist()]
                if any(c == "family" for c in cells) and any(c == "type" for c in cells):
                    header_row = i
                    break
            if header_row is None:
                return -1

            headers = [norm_text_lc(x).replace("–","-").replace("—","-").replace("‑","-").replace("−","-") for x in df0.iloc[header_row].tolist()]
            has_h = any(re.search(r"-\s*h\b", h) or h.endswith("-h") for h in headers)
            has_l = any(re.search(r"-\s*l\b", h) or h.endswith("-l") for h in headers)
            non_empty = sum(1 for h in headers if h.strip() != "")
            sc = 10 + (3 if has_h else 0) + (3 if has_l else 0) + min(5, non_empty // 3)
            return sc

        for sep in seps:
            try:
                raw_try = read_delimited_flexible(path, sep)
                sc = score_raw(raw_try)
                if sc > best_score:
                    best_score = sc
                    best_raw = raw_try
            except Exception:
                continue

        if best_raw is None:
            best_raw = read_delimited_flexible(path, seps[0])

        raw = best_raw.fillna("")

    raw = raw.loc[:, (raw.astype(str).apply(lambda col: (col.str.strip() != "").any())).values]

    header_row = None
    for i in range(min(len(raw), 40)):
        cells = [norm_text_lc(x) for x in raw.iloc[i].tolist()]
        if any(c == "family" for c in cells) and any(c == "type" for c in cells):
            header_row = i
            break

    if header_row is None:
        for i in range(min(len(raw), 40)):
            cells = [norm_text(x) for x in raw.iloc[i].tolist()]
            if sum(1 for c in cells if c != "") >= 2:
                header_row = i
                break

    if header_row is None:
        raise ValueError("Could not detect header row in input file.")

    headers = [norm_text(x) for x in raw.iloc[header_row].tolist()]
    seen = defaultdict(int)
    fixed = []
    for j, h in enumerate(headers):
        hh = h if h else f"COL{j+1}"
        seen[hh] += 1
        if seen[hh] > 1:
            hh = f"{hh}__{seen[hh]}"
        fixed.append(hh)

    df = raw.iloc[header_row + 1 :].copy()
    df.columns = fixed
    df = df.fillna("")
    return df


def find_col(df: pd.DataFrame, *candidates: str) -> Optional[str]:
    cols = list(df.columns)
    cols_lc = [norm_text_lc(c) for c in cols]

    for cand in candidates:
        key = norm_text_lc(cand)
        for i, lc in enumerate(cols_lc):
            if lc == key:
                return cols[i]

    for cand in candidates:
        key = norm_text_lc(cand)
        for i, lc in enumerate(cols_lc):
            if key in lc:
                return cols[i]
    return None


CHANNEL_SPEC = {
    "L50": "L 50x50x6",
    "L65": "L 65x65x6",
    "L75": "L 75x75x9",
    "C125": "C 125x65x6",
    "C150": "C 150x75x9",
}



TYPE1_PIPE_MAP = {
    2:  ('1 1/2"', 'Sch.80'),
    3:  ('2"',      'Sch.40'),
    4:  ('3"',      'Sch.40'),
    6:  ('4"',      'Sch.40'),
    8:  ('6"',      'Sch.40'),
    10: ('8"',      'Sch.40'),
    12: ('8"',      'Sch.40'),
    14: ('10"',     'Sch.40'),
    16: ('10"',     'Sch.40'),
    18: ('12"',     'Sch.40'),
    20: ('12"',     'Sch.40'),
}

TYPE1_PLATE_BOLT_MAP = {

    2:  ('PL 150x150x9', 'PL 290x290x9', 'EB M16x140'),
    3:  ('PL 150x150x9', 'PL 290x290x9', 'EB M16x140'),
    4:  ('PL 150x150x9', 'PL 290x290x9', 'EB M16x140'),

    6:  ('PL 230x230x9', 'PL 370x370x9', 'EB M16x140'),
    8:  ('PL 230x230x9', 'PL 370x370x9', 'EB M16x140'),

    10: ('PL 330x330x16', 'PL 490x490x16', 'EB M20x170'),
    12: ('PL 330x330x16', 'PL 490x490x16', 'EB M20x170'),
    14: ('PL 330x330x16', 'PL 490x490x16', 'EB M20x170'),
    16: ('PL 330x330x16', 'PL 490x490x16', 'EB M20x170'),

    18: ('PL 380x380x16', 'PL 560x560x16', 'EB M22x180'),
    20: ('PL 380x380x16', 'PL 560x560x16', 'EB M22x180'),
}

def round_up_100(mm: float) -> int:
    return int(math.ceil(float(mm) / 100.0) * 100)

def parse_type1_variant(type_str: str) -> str:
    s = norm_text(type_str)
    if not s:
        return "1"
    parts = [p.strip() for p in s.split("-") if p.strip()]
    if any(p.upper() == "A" for p in parts):
        return "1A"
    if parts and parts[-1].upper().endswith("C"):
        return "1C"
    return "1"

def parse_type1_support_inch(type_str: str) -> Optional[int]:
    s = norm_text(type_str)
    if not s:
        return None
    m = re.search(r"-(\d+)\s*B", s, flags=re.IGNORECASE)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None

@dataclass
class Type1Group:
    support_inch: int
    variant: str
    support_count: int
    pipe_len_counts: Dict[int, int]

def compute_type1(df: pd.DataFrame) -> Dict[Tuple[int, str], Type1Group]:
    """Aggregate Type 1 supports into groups by (support_inch, variant)."""
    type_col = find_col(df, "Type")
    total_col = find_col(df, "1-H total") or find_col(df, "1H total") or find_col(df, "1-h total")
    if not type_col or not total_col:
        raise ValueError("Type 1 file must contain columns: Type and 1-H total")

    groups: Dict[Tuple[int, str], Type1Group] = {}

    for _, row in df.iterrows():
        tval = norm_text(row.get(type_col, ""))
        if not tval:
            continue
        support_inch = parse_type1_support_inch(tval)
        if support_inch is None:
            continue
        variant = parse_type1_variant(tval)

        total_v = parse_number(row.get(total_col))
        if total_v is None:
            continue
        pipe_len = round_up_100(total_v)

        key = (support_inch, variant)
        if key not in groups:
            groups[key] = Type1Group(
                support_inch=support_inch,
                variant=variant,
                support_count=0,
                pipe_len_counts=defaultdict(int),
            )
        g = groups[key]
        g.support_count += 1
        g.pipe_len_counts[pipe_len] += 1

    return groups

def build_sheet_for_type1(ws, groups: Dict[Tuple[int, str], Type1Group], master_by_size: Dict[str, MasterRec]):
    row = 1
    row = write_title(ws, "Type 1 PIPE SUP'T", row)
    row = write_header(ws, row)

    def sort_key(k):
        inch, variant = k
        v_order = {"1": 0, "1A": 1, "1C": 2}.get(variant, 9)
        return (inch, v_order)

    for (inch, variant) in sorted(groups.keys(), key=sort_key):
        g = groups[(inch, variant)]
        if g.support_count == 0:
            continue

        label = "1" if variant == "1" else ("1-A" if variant == "1A" else "1-C")
        row = write_group(ws, f"Material List [For {inch}\" Type {label} PIPE SUP'T]", row)

        pipe_info = TYPE1_PIPE_MAP.get(inch)
        plate_bolt = TYPE1_PLATE_BOLT_MAP.get(inch)
        if not pipe_info or not plate_bolt:
            ws.cell(row=row, column=3, value="**UNKNOWN TYPE 1 INCH SIZE**")
            ws.cell(row=row, column=4, value=str(inch))
            row += 2
            continue

        pipe_inch_txt, pipe_sched = pipe_info
        small_plate, big_plate, bolt_size = plate_bolt

        for L in sorted(g.pipe_len_counts.keys()):
            qty = g.pipe_len_counts[L]
            pipe_size = f"{pipe_inch_txt} {pipe_sched} L {L}"
            rec = master_by_size.get(norm_text_lc(pipe_size)) or missing_rec(pipe_size)
            write_item_row(ws, row, rec, qty)
            row += 1

        elbow_size = f'{inch}" Sch.40 (Half Saddle)'
        elbow_key = norm_text_lc(elbow_size)
        elbow_rec = master_by_size.get(elbow_key) or missing_rec(elbow_size)

        if variant == "1A":
            elbow_rec = MasterRec(
                item_no=elbow_rec.item_no,
                material="ASTM A240 304/304L",
                name=elbow_rec.name,
                size=elbow_rec.size,
                treatment="-",
                unit=elbow_rec.unit,
                unit_weight=elbow_rec.unit_weight,
                unit_surface=elbow_rec.unit_surface,
                remark=elbow_rec.remark,
                add_notes=elbow_rec.add_notes,
                add_notes_lc=elbow_rec.add_notes_lc,
            )

        write_item_row(ws, row, elbow_rec, g.support_count)
        row += 1

        small_qty = g.support_count if variant != "1C" else 0
        big_qty = g.support_count

        sp_rec = master_by_size.get(norm_text_lc(small_plate)) or missing_rec(small_plate)
        bp_rec = master_by_size.get(norm_text_lc(big_plate)) or missing_rec(big_plate)

        write_item_row(ws, row, sp_rec, small_qty)
        row += 1
        write_item_row(ws, row, bp_rec, big_qty)
        row += 1

        bolt_qty = g.support_count * 4 if variant != "1C" else 0
        b_rec = master_by_size.get(norm_text_lc(bolt_size)) or missing_rec(bolt_size)
        write_item_row(ws, row, b_rec, bolt_qty)
        row += 2
def detect_file_type(df: pd.DataFrame) -> str:
    if find_col(df, "1-H total") or find_col(df, "1H total") or find_col(df, "1-h total"):
        return "1"

    type_col = find_col(df, "Type")
    if type_col:
        for v in df[type_col].tolist():
            s = norm_text(v)
            if not s:
                continue
            su = s.upper()

            if su.startswith("01-") or re.search(r"\b01-\d+\s*B", su):
                return "1"

            if su.startswith("66-"):
                return "66"
            if su.startswith("52-"):
                return "52"
            if su.startswith("54A"):
                return "54A"
            if su.startswith("54B"):
                return "54B"
            if su.startswith("23-"):
                return "23"
            if su.startswith("30-"):
                return "30"
            if su.startswith("31-"):
                return "31"
            if su.startswith("32-"):
                return "32"
            if su.startswith("33-"):
                return "33"
            if su.startswith("35-"):
                return "35"
    for k in CHANNEL_SPEC.keys():
        if find_col(df, k):
            return "STANDARD"
    return "STANDARD"


def extract_hl_cols(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    h_col = None
    l_col = None
    for c in df.columns:
        lc = norm_text_lc(c).replace("–","-").replace("—","-").replace("‑","-").replace("−","-")
        if h_col is None and (re.search(r"-\s*h\b", lc) or lc.endswith("-h")):
            h_col = c
        if l_col is None and (re.search(r"-\s*l\b", lc) or lc.endswith("-l")):
            l_col = c
    if not h_col:
        h_col = find_col(df, "H", "Height")
    if not l_col:
        l_col = find_col(df, "L", "Width", "Length")
    return h_col, l_col

def compute_standard(df: pd.DataFrame) -> Dict[str, Dict[int, int]]:
    h_col, l_col = extract_hl_cols(df)
    if not h_col or not l_col:
        raise ValueError("Could not find -H and -L columns in this file (after header auto-detection).")

    code_cols = {code: find_col(df, code) for code in CHANNEL_SPEC.keys()}
    counts_by_code: Dict[str, Dict[int, int]] = {k: defaultdict(int) for k in CHANNEL_SPEC.keys()}

    for _, row in df.iterrows():
        chosen_code = None
        for code, col in code_cols.items():
            if not col:
                continue
            if parse_yes(row[col]):
                chosen_code = code
                break
        if not chosen_code:
            continue

        def to_int(v) -> int:
            s = norm_text(v)
            if s == "":
                return 0
            try:
                return int(float(s))
            except Exception:
                return 0

        h0 = to_int(row[h_col])
        l0 = to_int(row[l_col])
        h = round_up_100(h0)
        l = round_up_100(l0)

        if h > 0:
            counts_by_code[chosen_code][h] += 1
        if l > 0:
            counts_by_code[chosen_code][l] += 1

    return counts_by_code

_A_LEN_RE = re.compile(r"\(\s*A\s*([0-9]+)\s*\)", re.IGNORECASE)

def extract_A_length_mm(type_str: str) -> Optional[int]:
    m = _A_LEN_RE.search(type_str)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None

def compute_type66(df: pd.DataFrame, master_by_size: Dict[str, MasterRec], master_recs: List[MasterRec]) -> Dict[str, Dict[str, int]]:
    type_col = find_col(df, "Type")
    if not type_col:
        raise ValueError("No Type column found for Type 66 rules.")
    groups: Dict[str, List[str]] = defaultdict(list)
    for v in df[type_col].tolist():
        s = norm_text(v)
        if s.upper().startswith("66-"):
            tok, _ = parse_inch_token(s)
            if tok:
                groups[tok].append(s)

    out: Dict[str, Dict[str, int]] = {}
    for tok, typestrs in groups.items():
        inch_val = inch_token_to_float(tok)
        band = inch_band(inch_val)
        n = len(typestrs)
        bucket: Dict[str, int] = defaultdict(int)

        pad = pick_padding_plate(master_recs, tok)
        if pad:
            bucket[norm_text_lc(pad.size)] += n * 1

        if band in {"LT10", "10_14"}:
            hch = pick_h_channel(master_recs, inch_val)
            if hch:
                bucket[norm_text_lc(hch.size)] += int(math.ceil(n / 2.0))

        if band == "10_14":
            rp = pick_reinforcement_plate(master_recs, tok)
            if rp:
                bucket[norm_text_lc(rp.size)] += n * 4

        if band == "GE16":
            for p in pick_pipe_shoe_plates(master_recs, inch_val):
                bucket[norm_text_lc(p.size)] += n * 1
            rp = pick_reinforcement_plate(master_recs, tok)
            if rp:
                bucket[norm_text_lc(rp.size)] += n * 4

        out[tok] = bucket
    return out

def compute_type52(df: pd.DataFrame, master_by_size: Dict[str, MasterRec], master_recs: List[MasterRec]) -> Dict[str, Dict[str, int]]:
    type_col = find_col(df, "Type")
    if not type_col:
        raise ValueError("No Type column found for Type 52 rules.")
    groups: Dict[str, List[str]] = defaultdict(list)
    for v in df[type_col].tolist():
        s = norm_text(v)
        if s.upper().startswith("52-"):
            tok, _ = parse_inch_token(s)
            if tok:
                groups[tok].append(s)

    out: Dict[str, Dict[str, int]] = {}
    small_rp = pick_small_reinforcement_plate(master_recs)

    for tok, typestrs in groups.items():
        inch_val = inch_token_to_float(tok)
        band = inch_band(inch_val)
        n = len(typestrs)
        bucket: Dict[str, int] = defaultdict(int)

        a_counts: Dict[int, int] = defaultdict(int)
        for t in typestrs:
            a_len = extract_A_length_mm(t) or 150
            a_counts[a_len] += 1
        for a_len, cnt in a_counts.items():
            fa = pick_forming_angle(master_by_size, a_len)
            if fa:
                bucket[norm_text_lc(fa.size)] += cnt * 2

        pad = pick_padding_plate(master_recs, tok)
        if pad:
            bucket[norm_text_lc(pad.size)] += n * 1

        if band in {"LT10", "10_14"}:
            hch = pick_h_channel(master_recs, inch_val)
            if hch:
                bucket[norm_text_lc(hch.size)] += int(math.ceil(n / 2.0))

        rp = pick_reinforcement_plate(master_recs, tok)

        if band == "10_14":
            if small_rp:
                bucket[norm_text_lc(small_rp.size)] += n * 4
            if rp:
                bucket[norm_text_lc(rp.size)] += n * 4

        if band == "GE16":
            for p in pick_pipe_shoe_plates(master_recs, inch_val):
                bucket[norm_text_lc(p.size)] += n * 1
            if small_rp:
                bucket[norm_text_lc(small_rp.size)] += n * 4
            if rp:
                bucket[norm_text_lc(rp.size)] += n * 4

        out[tok] = bucket

    return out

def compute_type54(df: pd.DataFrame, which: str, master_by_size: Dict[str, MasterRec], master_recs: List[MasterRec]) -> Dict[str, Dict[str, int]]:
    assert which in {"54A", "54B"}
    type_col = find_col(df, "Type")
    if not type_col:
        raise ValueError("No Type column found for Type 54 rules.")
    groups: Dict[str, List[str]] = defaultdict(list)
    for v in df[type_col].tolist():
        s = norm_text(v)
        su = s.upper()
        if which == "54A":
            ok = su.startswith("54A")
        else:
            ok = su.startswith("54B")
        if not ok:
            continue
        tok, _ = parse_inch_token(s)
        if tok:
            groups[tok].append(s)

    out: Dict[str, Dict[str, int]] = {}
    small_rp = pick_small_reinforcement_plate(master_recs)

    for tok, typestrs in groups.items():
        inch_val = inch_token_to_float(tok)
        band = inch_band(inch_val)
        n = len(typestrs)
        bucket: Dict[str, int] = defaultdict(int)

        clamp = pick_pipe_clamp(master_recs, tok)
        gasket = pick_gasket(master_recs, tok)
        bolt = pick_hex_bolt_set(master_recs, inch_val)

        if clamp:
            bucket[norm_text_lc(clamp.size)] += n * 4
        if gasket:
            bucket[norm_text_lc(gasket.size)] += n * 2
        if bolt:
            bucket[norm_text_lc(bolt.size)] += n * 4

        if which == "54A":
            a_counts: Dict[int, int] = defaultdict(int)
            for t in typestrs:
                a_len = extract_A_length_mm(t) or 150
                a_counts[a_len] += 1
            for a_len, cnt in a_counts.items():
                fa = pick_forming_angle(master_by_size, a_len)
                if fa:
                    bucket[norm_text_lc(fa.size)] += cnt * 2

        if band in {"LT10", "10_14"}:
            hch = pick_h_channel(master_recs, inch_val)
            if hch:
                bucket[norm_text_lc(hch.size)] += int(math.ceil(n / 2.0))

        rp = pick_reinforcement_plate(master_recs, tok)

        if band == "10_14":
            if which == "54A":
                if small_rp:
                    bucket[norm_text_lc(small_rp.size)] += n * 4
                if rp:
                    bucket[norm_text_lc(rp.size)] += n * 4
            else:
                if rp:
                    bucket[norm_text_lc(rp.size)] += n * 4

        if band == "GE16":
            for p in pick_pipe_shoe_plates(master_recs, inch_val):
                bucket[norm_text_lc(p.size)] += n * 1

            if which == "54A":
                if small_rp:
                    bucket[norm_text_lc(small_rp.size)] += n * 4
                if rp:
                    bucket[norm_text_lc(rp.size)] += n * 4
            else:
                if small_rp:
                    bucket[norm_text_lc(small_rp.size)] += n * 4
                if rp:
                    bucket[norm_text_lc(rp.size)] += n * 4

        out[tok] = bucket

    return out


BORDER_THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
FILL_TITLE = PatternFill("solid", fgColor="D9D9D9")
FILL_GROUP = PatternFill("solid", fgColor="BFBFBF")
FILL_HEADER = PatternFill("solid", fgColor="E7E6E6")
FONT_BOLD = Font(bold=True)

ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT_WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTER_TOP_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)

OUTPUT_HEADERS = [
    "ITEM NO.", "MATERIAL", "NAME", "SIZE", "TREATMENT", "UNIT", "Q'TY",
    "UNIT WEIGHT (KG/PCS)", "TOTAL WEIGHT (KG)",
    "UNIT SURFACE AREA (M2)", "TOTAL SURFACE (M2)",
    "REMARK", "ADD. NOTES"
]

def set_col_widths(ws):
    widths = {
        "A": 10, "B": 14, "C": 26, "D": 28, "E": 22, "F": 10, "G": 10,
        "H": 18, "I": 16, "J": 20, "K": 18, "L": 18, "M": 46,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def style_row(ws, row_idx: int, fill=None, bold=False):
    for c in range(1, 13 + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER_WRAP
        if fill:
            cell.fill = fill
        if bold:
            cell.font = FONT_BOLD

def write_title(ws, title: str, row_idx: int) -> int:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.fill = FILL_TITLE
    cell.font = Font(bold=True, size=12)
    cell.alignment = ALIGN_LEFT_WRAP_TOP
    style_row(ws, row_idx, fill=FILL_TITLE, bold=True)
    ws.row_dimensions[row_idx].height = 22
    return row_idx + 1

def write_header(ws, row_idx: int) -> int:
    for i, h in enumerate(OUTPUT_HEADERS, start=1):
        ws.cell(row=row_idx, column=i, value=h)
    style_row(ws, row_idx, fill=FILL_HEADER, bold=True)
    ws.row_dimensions[row_idx].height = 30
    return row_idx + 1

def write_group(ws, label: str, row_idx: int) -> int:
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
    cell = ws.cell(row=row_idx, column=1, value=label)
    cell.fill = FILL_GROUP
    cell.font = FONT_BOLD
    cell.alignment = ALIGN_LEFT_WRAP_TOP
    style_row(ws, row_idx, fill=FILL_GROUP, bold=True)
    ws.row_dimensions[row_idx].height = 20
    return row_idx + 1

def estimate_lines(text: str, chars_per_line: int) -> int:
    if not text:
        return 1
    parts = text.split("\n")
    lines = 0
    for p in parts:
        p = p.strip()
        if not p:
            lines += 1
        else:
            lines += max(1, int(math.ceil(len(p) / chars_per_line)))
    return max(1, lines)

def set_item_row_height(ws, row_idx: int, remark: str, add_notes: str):
    r_lines = estimate_lines(remark, 16)
    a_lines = estimate_lines(add_notes, 34)
    lines = max(r_lines, a_lines, 1)

    height = min(18 * lines, 180)
    ws.row_dimensions[row_idx].height = max(18, height)


def missing_rec(size_display: str) -> MasterRec:
    return MasterRec(
        item_no="**",
        material="**",
        name="**MISSING IN MASTER**",
        size=size_display,
        treatment="**",
        unit="**",
        unit_weight=None,
        unit_surface=None,
        remark="**",
        add_notes="**",
        add_notes_lc="",
    )

def write_item_row(ws, row_idx: int, rec: MasterRec, qty: int):
    ws.cell(row=row_idx, column=1, value=rec.item_no or "**")
    ws.cell(row=row_idx, column=2, value=rec.material or "**")
    ws.cell(row=row_idx, column=3, value=rec.name or "**")
    ws.cell(row=row_idx, column=4, value=rec.size or "**")
    ws.cell(row=row_idx, column=5, value=rec.treatment or "**")
    ws.cell(row=row_idx, column=6, value=rec.unit or "**")
    ws.cell(row=row_idx, column=7, value=int(qty))

    if rec.unit_weight is None:
        ws.cell(row=row_idx, column=8, value="**")
    else:
        ws.cell(row=row_idx, column=8, value=float(rec.unit_weight))

    ws.cell(row=row_idx, column=9, value=f_total_weight(row_idx))

    if rec.unit_surface is None:
        ws.cell(row=row_idx, column=10, value="**")
    else:
        ws.cell(row=row_idx, column=10, value=float(rec.unit_surface))

    ws.cell(row=row_idx, column=11, value=f_total_surface(row_idx))

    ws.cell(row=row_idx, column=12, value=rec.remark if rec.remark else "**")
    ws.cell(row=row_idx, column=13, value=rec.add_notes if rec.add_notes else "**")

    for c in range(1, 14):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER_TOP_WRAP

    ws.cell(row=row_idx, column=12).alignment = ALIGN_LEFT_WRAP_TOP
    ws.cell(row=row_idx, column=13).alignment = ALIGN_LEFT_WRAP_TOP

    set_item_row_height(ws, row_idx, ws.cell(row=row_idx, column=12).value, ws.cell(row=row_idx, column=13).value)


Sig = Tuple[str, str, str, str, str]

def rec_signature(rec: MasterRec) -> Sig:
    return (
        norm_text_lc(rec.size),
        norm_text_lc(rec.material),
        norm_text_lc(rec.treatment),
        norm_text_lc(rec.name),
        norm_text_lc(rec.unit),
    )

def add_to_totals(totals: Dict[Sig, Tuple[MasterRec, int]], rec: MasterRec, qty: int):
    q = int(qty or 0)
    sig = rec_signature(rec)
    if sig in totals:
        old_rec, old_q = totals[sig]
        totals[sig] = (old_rec, old_q + q)
    else:
        totals[sig] = (rec, q)

def build_summary_sheet(wb_out: Workbook, master_recs: List[MasterRec], totals: Dict[Sig, Tuple[MasterRec, int]]):
    name = "總表"
    if name in wb_out.sheetnames:
        i = 1
        while f"{name}_{i}" in wb_out.sheetnames:
            i += 1
        name = f"{name}_{i}"
    ws = wb_out.create_sheet(name, 0)
    set_col_widths(ws)

    row = 1
    row = write_title(ws, "總表", row)
    row = write_header(ws, row)

    master_sigs = set()
    for rec in master_recs:
        sig = rec_signature(rec)
        master_sigs.add(sig)
        qty = totals.get(sig, (rec, 0))[1]
        write_item_row(ws, row, rec, qty)
        row += 1

    extras = [(rec, qty) for sig, (rec, qty) in totals.items() if sig not in master_sigs and int(qty) != 0]
    if extras:
        row += 1
        row = write_group(ws, "EXTRA (Not in Master)", row)
        extras.sort(key=lambda x: (norm_text_lc(x[0].name), norm_text_lc(x[0].size), norm_text_lc(x[0].material)))
        for rec, qty in extras:
            write_item_row(ws, row, rec, qty)
            row += 1
        row += 1

    first_data = 3
    last_data = row - 1
    if last_data >= first_data:
        for c in range(1, 14):
            ws.cell(row=row, column=c, value="**")
        ws.cell(row=row, column=3, value="TOTAL")
        ws.cell(row=row, column=7, value=f"=SUM(G{first_data}:G{last_data})")
        ws.cell(row=row, column=9, value=f"=SUM(I{first_data}:I{last_data})")
        ws.cell(row=row, column=11, value=f"=SUM(K{first_data}:K{last_data})")
        style_row(ws, row, fill=FILL_GROUP, bold=True)
        ws.row_dimensions[row].height = 20


def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", name).strip()
    return (name or "Sheet")[:31]


def build_sheet_for_standard(ws, df: pd.DataFrame, file_label: str, master_by_size: Dict[str, MasterRec]):
    row = 1
    row = write_title(ws, file_label, row)
    row = write_header(ws, row)

    counts_by_code = compute_standard(df)

    for code, counts in counts_by_code.items():
        if not counts:
            continue
        row = write_group(ws, f"Material List [ For {code} {file_label} ]", row)
        spec = CHANNEL_SPEC.get(code, code)
        max_len = max(counts.keys()) if counts else 0
        for length in range(100, max_len + 1, 100):
            size = f"{spec}x{length}"
            rec = master_by_size.get(norm_text_lc(size)) or missing_rec(size)
            qty = counts.get(length, 0)
            write_item_row(ws, row, rec, qty)
            row += 1
        row += 1

def build_sheet_for_special(ws, groups: Dict[str, Dict[str, int]], file_label: str, master_by_size: Dict[str, MasterRec]):
    row = 1
    row = write_title(ws, file_label, row)
    row = write_header(ws, row)

    def inch_sort_key(tok: str):
        v = inch_token_to_float(tok)
        return (1e9, tok) if math.isnan(v) else (v, tok)

    for tok in sorted(groups.keys(), key=inch_sort_key):
        bucket = groups[tok]
        if not bucket:
            continue
        row = write_group(ws, f"Material List [ For {tok}\" {file_label} ]", row)

        items = []
        for size_key, qty in bucket.items():
            if qty == 0:
                continue
            rec = master_by_size.get(size_key) or missing_rec(size_key)
            items.append((rec, qty))

        def item_no_key(rec: MasterRec):
            try:
                return (int(str(rec.item_no).strip()), rec.size)
            except Exception:
                return (10**9, rec.size)

        items.sort(key=lambda x: item_no_key(x[0]))

        for rec, qty in items:
            write_item_row(ws, row, rec, qty)
            row += 1

        row += 1


def process_one_file(
    path: str,
    wb_out: Workbook,
    master_by_size: Dict[str, MasterRec],
    master_recs: List[MasterRec],
    totals: Dict[Sig, Tuple[MasterRec, int]],
):
    df = read_input_table(path)
    ftype = detect_file_type(df)
    base = os.path.splitext(os.path.basename(path))[0]
    ws = wb_out.create_sheet(sanitize_sheet_name(base))
    set_col_widths(ws)

    if ftype == "1":
        groups = compute_type1(df)
        build_sheet_for_type1(ws, groups, master_by_size)

        for (inch, variant), g in groups.items():
            if g.support_count <= 0:
                continue

            pipe_info = TYPE1_PIPE_MAP.get(inch)
            plate_bolt = TYPE1_PLATE_BOLT_MAP.get(inch)
            if not pipe_info or not plate_bolt:
                continue

            pipe_inch_txt, pipe_sched = pipe_info
            small_plate, big_plate, bolt_size = plate_bolt

            for L, qty in g.pipe_len_counts.items():
                if qty <= 0:
                    continue
                pipe_size = f"{pipe_inch_txt} {pipe_sched} L {L}"
                rec = master_by_size.get(norm_text_lc(pipe_size)) or missing_rec(pipe_size)
                add_to_totals(totals, rec, qty)

            elbow_size = f'{inch}" Sch.40 (Half Saddle)'
            elbow_key = norm_text_lc(elbow_size)
            elbow_rec = master_by_size.get(elbow_key) or missing_rec(elbow_size)

            if variant == "1A":
                elbow_rec = MasterRec(
                    item_no=elbow_rec.item_no,
                    material="ASTM A240 304/304L",
                    name=elbow_rec.name,
                    size=elbow_rec.size,
                    treatment="-",
                    unit=elbow_rec.unit,
                    unit_weight=elbow_rec.unit_weight,
                    unit_surface=elbow_rec.unit_surface,
                    remark=elbow_rec.remark,
                    add_notes=elbow_rec.add_notes,
                    add_notes_lc=elbow_rec.add_notes_lc,
                )

            add_to_totals(totals, elbow_rec, g.support_count)

            small_qty = g.support_count if variant != "1C" else 0
            big_qty = g.support_count

            sp_rec = master_by_size.get(norm_text_lc(small_plate)) or missing_rec(small_plate)
            bp_rec = master_by_size.get(norm_text_lc(big_plate)) or missing_rec(big_plate)

            if small_qty:
                add_to_totals(totals, sp_rec, small_qty)
            add_to_totals(totals, bp_rec, big_qty)

            bolt_qty = g.support_count * 4 if variant != "1C" else 0
            if bolt_qty:
                b_rec = master_by_size.get(norm_text_lc(bolt_size)) or missing_rec(bolt_size)
                add_to_totals(totals, b_rec, bolt_qty)

    elif ftype == "66":
        groups = compute_type66(df, master_by_size, master_recs)
        build_sheet_for_special(ws, groups, "Type 66 PIPE SUP'T", master_by_size)
        for bucket in groups.values():
            for size_key, qty in bucket.items():
                if qty <= 0:
                    continue
                rec = master_by_size.get(size_key) or missing_rec(size_key)
                add_to_totals(totals, rec, qty)

    elif ftype == "52":
        groups = compute_type52(df, master_by_size, master_recs)
        build_sheet_for_special(ws, groups, "Type 52 PIPE SUP'T", master_by_size)
        for bucket in groups.values():
            for size_key, qty in bucket.items():
                if qty <= 0:
                    continue
                rec = master_by_size.get(size_key) or missing_rec(size_key)
                add_to_totals(totals, rec, qty)

    elif ftype == "54A":
        groups = compute_type54(df, "54A", master_by_size, master_recs)
        build_sheet_for_special(ws, groups, "Type 54A PIPE SUP'T", master_by_size)
        for bucket in groups.values():
            for size_key, qty in bucket.items():
                if qty <= 0:
                    continue
                rec = master_by_size.get(size_key) or missing_rec(size_key)
                add_to_totals(totals, rec, qty)

    elif ftype == "54B":
        groups = compute_type54(df, "54B", master_by_size, master_recs)
        build_sheet_for_special(ws, groups, "Type 54B PIPE SUP'T", master_by_size)
        for bucket in groups.values():
            for size_key, qty in bucket.items():
                if qty <= 0:
                    continue
                rec = master_by_size.get(size_key) or missing_rec(size_key)
                add_to_totals(totals, rec, qty)

    else:
        counts_by_code = compute_standard(df)
        build_sheet_for_standard(ws, df, base, master_by_size)

        for code, counts in counts_by_code.items():
            if not counts:
                continue
            spec = CHANNEL_SPEC.get(code, code)
            for length, qty in counts.items():
                if qty <= 0:
                    continue
                size = f"{spec}x{length}"
                rec = master_by_size.get(norm_text_lc(size)) or missing_rec(size)
                add_to_totals(totals, rec, qty)
def run_cli(master_path: str, inputs: List[str], output_path: str):
    master_by_size, master_recs = load_master(master_path)

    wb_out = Workbook()
    if wb_out.worksheets:
        wb_out.remove(wb_out.worksheets[0])

    totals: Dict[Sig, Tuple[MasterRec, int]] = {}

    for p in inputs:
        try:
            process_one_file(p, wb_out, master_by_size, master_recs, totals)
        except Exception as e:
            ws = wb_out.create_sheet(sanitize_sheet_name(os.path.splitext(os.path.basename(p))[0] + "_ERROR"))
            ws["A1"].value = f"ERROR processing file: {p}"
            ws["A2"].value = str(e)

    try:
        build_summary_sheet(wb_out, master_recs, totals)
    except Exception as e:
        ws = wb_out.create_sheet("總表_ERROR", 0)
        ws["A1"].value = "ERROR building summary sheet (總表)"
        ws["A2"].value = str(e)

    wb_out.calculation.fullCalcOnLoad = True
    wb_out.save(output_path)
def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    messagebox.showinfo(
        "MTO Generator",
        "Step 1) Select Master.xlsx (reference)\n"
        "Step 2) Select one or more Type files\n"
        "Step 3) Save output workbook",
    )

    master_path = filedialog.askopenfilename(
        title="Select Master.xlsx (reference)",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")],
    )
    if not master_path:
        return

    inputs: List[str] = []
    while True:
        new_files = filedialog.askopenfilenames(
            title="Select Type files (you can select many)",
            filetypes=[
                ("All supported", "*.csv *.tsv *.txt *.xlsx *.xlsm *.xls"),
                ("CSV/TSV", "*.csv *.tsv *.txt"),
                ("Excel", "*.xlsx *.xlsm *.xls"),
                ("All files", "*.*"),
            ],
        )
        if not new_files:
            break
        inputs.extend(list(new_files))
        if not messagebox.askyesno("Add more?", "Add more input files?"):
            break
    if not inputs:
        return

    out_path = filedialog.asksaveasfilename(
        title="Save output workbook",
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
        initialfile="MTO_Output.xlsx",
    )
    if not out_path:
        return

    try:
        run_cli(master_path, inputs, out_path)
        messagebox.showinfo("Done", f"Saved:\n{out_path}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Generate MTO workbook from Type files using Master.xlsx")
    ap.add_argument("--master", default="", help="Path to Master.xlsx")
    ap.add_argument("--out", default="MTO_Output.xlsx", help="Output xlsx path")
    ap.add_argument("inputs", nargs="*", help="Input Type files (csv/tsv/xlsx)")
    args = ap.parse_args()

    if args.master and args.inputs:
        run_cli(args.master, args.inputs, args.out)
    else:
        run_gui()
